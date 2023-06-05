# region Import
import os
import psycopg2
import openpyxl
import shutil
from dotenv import find_dotenv, load_dotenv

# endregion

load_dotenv(find_dotenv())


def main():
    # add_certificates(get_cert_info_from_user("cert_upload_template.xlsx"))
    add_products(get_product_info_from_user('products_upload_template.xlsx'))
    # column_name_query('information_schema.columns','certificates','column_name')
    # pass
    # table_data_query('information_schema.columns', 'certificates', 'column_name')


# region Query Helpers
def universal_query(table_name, *args):
    args = ','.join(args)
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
        table_len = cursor.fetchone()[0]
        cursor.execute(f"SELECT {args} FROM {table_name}")
        query_data = cursor.fetchmany(table_len)
        return query_data
    except Exception as _ex:
        print(f"[INFO] ERROR while working with data base {_ex}")
    finally:
        cursor.close()
        connect.close()


def table_data_query(table_name, where_filter, *args):
    args = ','.join(args)
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.execute(f"SELECT COUNT({args}) FROM {table_name}"
                       f" WHERE table_name = '{where_filter}'")
        table_len = cursor.fetchone()[0]
        cursor.execute(f"SELECT {args} FROM {table_name} "
                       f"WHERE table_name = '{where_filter}'"
                       f"ORDER BY ordinal_position ASC")
        column_name = cursor.fetchmany(table_len)
        column_name = [x[0] for x in column_name]
        return column_name
    except Exception as _ex:
        print(f"[INFO] ERROR while working with data base {_ex}")
    finally:
        cursor.close()
        connect.close()


# endregion

# region Levinstain algo
def levinstain_algo(i, j, s1, s2, matrix):
    if i == 0 and j == 0:
        return 0
    if i == 0 and j > 0:
        return j
    if j == 0 and i > 0:
        return i
    else:
        if s1[i - 1] == s2[j - 1]:
            m = 0
        else:
            m = 1
    return min(matrix[i][j - 1] + 1, matrix[i - 1][j] + 1, matrix[i - 1][j - 1] + m)


def calculate_levinstain_distance(s1, s2):
    str_1 = len(s1)
    str_2 = len(s2)
    matrix = [[0 for x in range(str_2 + 1)] for j in range(str_1 + 1)]
    for i in range(str_1 + 1):
        for j in range(str_2 + 1):
            matrix[i][j] = levinstain_algo(i, j, s1, s2, matrix)
    return matrix[str_1][str_2]


# endregion

# region Get Editorial number
def get_editorial_num(input_string):
    if 2 <= len(input_string) <= 6:
        editorial_number = 2
    else:
        editorial_number = 3
    return editorial_number


# endregion


# region Make replacment
def make_replacement(existing_pkey, user_data, table_name, where_filter):
    columns_name_list = table_data_query('information_schema.columns', table_name, 'column_name')
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        for i in range(len(existing_pkey)):
            for j in range(1, len(columns_name_list)):
                cursor.execute(f" UPDATE {table_name}"
                               f" SET {columns_name_list[j]} = %s"
                               f" WHERE {where_filter} = %s ", (user_data[i][j], user_data[i][0]))
    except Exception as _ex:
        print(f"[INFO] ERROR while working with data base {_ex}")
    finally:
        cursor.close()
        connect.close()


# endregion


# region Get_info_from_user
def get_cert_info_from_user(file_path):
    if os.path.isfile(file_path):
        certification_data_list = []
        book = openpyxl.open(file_path, read_only=True, data_only=True)
        sheet = book.active
        for row in range(3, sheet.max_row + 1):
            tmp_list = []
            for column in range(0, 5):
                tmp_list.append(sheet[row][column].value)
            if None in tmp_list:
                print("Columns can't be empty")
                return None
            else:
                certification_data_list.append(tmp_list)
        return certification_data_list
    else:
        raise ValueError("The function argument must be a file")


def get_product_info_from_user(file_path):
    def make_manufacrurer_dict(query):
        manufacturer_dict = {}
        for x in query:
            if x[1] not in manufacturer_dict:
                manufacturer_dict[x[1]] = x[0]
            else:
                pass
        return manufacturer_dict

    def make_possible_change_dict(dict_to_compression, user_data_list):
        possible_change_list = []
        for i in range(len(user_data_list)):
            tmp_list = []
            editorial_number = get_editorial_num(user_data_list[i][3])
            for key in dict_to_compression.keys():
                if editorial_number >= calculate_levinstain_distance(key, user_data_list[i][3].upper()):
                    tmp_list.append(key.upper())
                else:
                    continue
            tmp_list.append((user_data_list[i][3]))
            possible_change_list.append(tmp_list)
        possible_change_dict = {}
        for j in range(len(possible_change_list)):
            possible_change_dict[possible_change_list[j][-1]] = possible_change_list[j][:-1]
        return possible_change_dict

    if os.path.isfile(file_path):
        products_data_list = []
        book = openpyxl.open(file_path, read_only=True, data_only=True)
        sheet = book.active
        for row in range(2, sheet.max_row + 1):
            tmp_list = []
            for column in range(0, 7):
                tmp_list.append(sheet[row][column].value)
            products_data_list.append(tmp_list)
        manufacturer_dict = make_manufacrurer_dict(universal_query('manufacturers', '*'))
        print(manufacturer_dict)
        print(products_data_list)
        approved_data_list = []
        data_list_to_check = []
        for i in range(len(products_data_list)):
            if products_data_list[i][3].upper() in manufacturer_dict:
                products_data_list[i][3] = manufacturer_dict[products_data_list[i][3].upper()]
                approved_data_list.append(products_data_list[i])
            else:
                data_list_to_check.append(products_data_list[i])
        possible_change = make_possible_change_dict(manufacturer_dict, data_list_to_check)
        if possible_change:
            offer_to_user_to_check_manufacturers = input(str(f"В вашем файле присуствуют поставщики которых"
                                                             f" нет в списке поставщиков в базеданых\n"
                                                             f"Если хотите проверить данные, нажмите 'y' и программа"
                                                             f" отправит вам файл с возможными заменами\n"
                                                             f"если вы уверены в своем выборе нажмите любую"
                                                             f" другую кнопку и программа добавит в"
                                                             f" таблицу новых поставщиков"))
            if offer_to_user_to_check_manufacturers != 'y':
                try:
                    connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                               password=os.getenv('password'), host=os.getenv('host'))
                    connect.autocommit = True
                    cursor = connect.cursor()
                    print("Postgre SQL successfully conected")
                    for j in range(len(data_list_to_check)):
                        cursor.execute(f"INSERT INTO manufacturers (manufacturer_name) "
                                       f"VALUES ('{data_list_to_check[j][3].upper()}')")
                    manufacturer_dict = make_manufacrurer_dict(universal_query('manufacturers', '*'))
                    for i in range(len(products_data_list)):
                        if str(products_data_list[i][3]).upper() in manufacturer_dict.keys():
                            products_data_list[i][3] = manufacturer_dict[products_data_list[i][3].upper()]
                        else:
                            pass
                except Exception as _ex:
                    print(f"[INFO] ERROR while working with data base {_ex}")
                finally:
                    cursor.close()
                    connect.close()
                    print("Postgre SQL Connection closed")
                    return products_data_list
            else:
                shutil.copy(rf"{file_path}", rf"possible_changes.xlsx")
                book = openpyxl.open(rf"possible_changes.xlsx", read_only=False, data_only=True)
                sheet = book.active
                for row in range(2, sheet.max_row + 1):
                    for column in range(0, 7):
                        if sheet[row][3].value in possible_change.keys():
                            sheet[row][3].value = f" Список замен для производителя {sheet[row][3].value}" \
                                                  f" следующий:\n" \
                                                  f"{possible_change[sheet[row][3].value]}"
                        else:
                            index = sheet[row]
                            print(index)
                book.save(rf"possible_changes.xlsx")
                book.close()
                return approved_data_list
        else:
            return products_data_list

    else:
        raise ValueError("The function argument must be a file")


# endregion


# region add_info_to_table
def add_certificates(cert_data_from_user):
    existing_pkey = universal_query('certificates', 'certificate_id')
    existing_pkey = [x[0] for x in existing_pkey]
    unique_user_data = []
    duplicate_user_data = []

    for i in range(len(cert_data_from_user)):
        if cert_data_from_user[i][0] in existing_pkey:
            duplicate_user_data.append(cert_data_from_user[i])
        else:
            unique_user_data.append(cert_data_from_user[i])
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        if duplicate_user_data:
            if unique_user_data != '':
                cursor.executemany(f""" INSERT INTO certificates (certificate_id,certificate_number,certificate_type_id,
                                                        start_date,end_date)
                                                        VALUES
                                                        (%s,%s,%s,%s,%s) """, unique_user_data)
            else:
                pass
            add_duplicates_question = str(input(f'В базе данные обнаружены дубликаты данных'
                                                f' по следующим ID {duplicate_user_data}. Заменить данные? y/n?'))
            if add_duplicates_question == 'y':
                existing_pkey = duplicate_user_data
                make_replacement(existing_pkey, duplicate_user_data, 'certificates', 'certificate_id')
            else:
                pass
        else:
            cursor.executemany(f""" INSERT INTO certificates (certificate_id,certificate_number,certificate_type_id,
                                        start_date,end_date)
                                        VALUES
                                        (%s,%s,%s,%s,%s) """, cert_data_from_user)
    except Exception as _ex:
        print(f"[INFO] ERROR while working with data base {_ex}")
    finally:
        cursor.close()
        connect.close()
        print("Postgre SQL Connection closed")


def add_products(products_data_from_user):
    existing_pkey = universal_query('bilight_products', 'product_id')
    existing_pkey = [x[0] for x in existing_pkey]
    unique_user_data = []
    duplicate_user_data = []

    for i in range(len(products_data_from_user)):
        if products_data_from_user[i][0] in existing_pkey:
            duplicate_user_data.append(products_data_from_user[i])
        else:
            unique_user_data.append(products_data_from_user[i])
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        if duplicate_user_data:
            if unique_user_data != '':
                cursor.executemany(f""" INSERT INTO bilight_products (product_id,order_code,article,
                                                                manufacturer_id,product_name,tnved_id,certificate_id)
                                                                VALUES
                                                                (%s,%s,%s,%s,%s,%s,%s) """, unique_user_data)
            else:
                pass
            add_duplicates_question = str(input(f'В базе данные обнаружены дубликаты данных'
                                                f' по следующим ID {duplicate_user_data}. Заменить данные? y/n?'))
            if add_duplicates_question == 'y':
                existing_pkey = duplicate_user_data
                make_replacement(existing_pkey, duplicate_user_data, 'bilight_products', 'product_id')
            else:
                pass
        else:
            cursor.executemany(f""" INSERT INTO bilight_products (product_id,order_code,article,
                                                                manufacturer_id,product_name,tnved_id,certificate_id)
                                                                VALUES
                                                                (%s,%s,%s,%s,%s,%s,%s) """, products_data_from_user)
    except Exception as _ex:
        print(f"[INFO] ERROR while working with data base {_ex}")
    finally:
        cursor.close()
        connect.close()
        print("Postgre SQL Connection closed")


# endregion


if __name__ == '__main__':
    main()
