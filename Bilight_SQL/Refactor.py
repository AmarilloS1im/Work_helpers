# region Import
import logging
import os
import psycopg2
import openpyxl
import shutil
from openpyxl.styles import Font, Alignment
from dotenv import find_dotenv, load_dotenv
import uuid
from loguru import logger

load_dotenv(find_dotenv())


# endregion



# region Levinstain algo
def levinstain_algo(i, j, s1, s2, matrix):
    if i == 0 and j == 0:
        return 0
    if i == 0 and j > 0:
        return j
    if j == 0 and i > 0:
        return i
    else:
        if s1[i - 1] == s2[j - 1]:
            m = 0
        else:
            m = 1
    return min(matrix[i][j - 1] + 1, matrix[i - 1][j] + 1, matrix[i - 1][j - 1] + m)


def calculate_levinstain_distance(s1, s2):
    str_1 = len(s1)
    str_2 = len(s2)
    matrix = [[0 for x in range(str_2 + 1)] for j in range(str_1 + 1)]
    for i in range(str_1 + 1):
        for j in range(str_2 + 1):
            matrix[i][j] = levinstain_algo(i, j, s1, s2, matrix)
    return matrix[str_1][str_2]


# endregion

# region GET FUNCTIONS
def universal_query(table_name, *args):
    args = ','.join(args)
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM {table_name}")
        table_len = cursor.fetchone()[0]
        cursor.execute(f"SELECT {args} FROM {table_name}")
        query_data = cursor.fetchmany(table_len)
        return query_data
    except Exception as _ex:
        logger.add('error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def get_column_name_list(table_name, where_filter, *args):
    args = ','.join(args)
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM {table_name}"
                       f" WHERE table_name = '{where_filter}'")
        table_len = cursor.fetchone()[0]
        cursor.execute(f"SELECT {args} FROM {table_name} "
                       f"WHERE table_name = '{where_filter}'"
                       f"ORDER BY ordinal_position ASC")
        column_name = cursor.fetchmany(table_len)
        column_name = [x[0] for x in column_name]
        return column_name
    except Exception as _ex:
        logger.add('error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def get_all_article_from_db():
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.execute(f"SELECT COUNT(*) FROM bilight_products")
        table_len = cursor.fetchone()[0]
        cursor.execute(f"SELECT product_id,order_code FROM bilight_products ")
        all_articles = cursor.fetchmany(table_len)
        return all_articles
    except Exception as _ex:
        logger.add('error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def get_id_by_article_query(article, needed_id_name):
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.execute(
            f"SELECT {needed_id_name} FROM bilight_products "
            f"WHERE product_id = '{article}' or order_code = '{article}'")
        needed_id = cursor.fetchone()[0]
        return needed_id
    except Exception as _ex:
        logger.add('error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def get_product_info_from_user(file_path):
    products_data_list = []
    book = openpyxl.open(file_path, read_only=True, data_only=True)
    sheet = book.active
    for row in range(2, sheet.max_row + 1):
        tmp_list = []
        for column in range(0, 7):
            tmp_list.append(sheet[row][column].value)
        products_data_list.append(tmp_list)
    book.close()
    return products_data_list


def get_cert_info_from_user(file_path):
    uuid = file_path.split('_')[0]
    certification_data_list = []
    book = openpyxl.open(file_path, read_only=True, data_only=True)
    sheet = book.active
    for row in range(3, sheet.max_row + 1):
        tmp_list = []
        for column in range(0, 5):
            tmp_list.append(sheet[row][column].value)
        if None in tmp_list:
            logger.add(f'{uuid}_error.log')
            logger.error("Columns can't be empty")
            logger.remove()
            book.close()
            return None
        else:
            certification_data_list.append(tmp_list)
    book.close()
    return certification_data_list


def get_editorial_num(input_string):
    if 2 <= len(input_string) <= 6:
        editorial_number = 2
    else:
        editorial_number = 3
    return editorial_number


# endregion

# region ADD FUNCTIONS
def add_unique_user_data(unique_data):
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        if unique_data != '':
            cursor.executemany(f""" INSERT INTO bilight_products (product_id,order_code,article,
                                                            manufacturer_id,product_name,tnved_id,certificate_id)
                                                            VALUES
                                                            (%s,%s,%s,%s,%s,%s,%s) """, unique_data)
        else:
            pass
    except Exception as _ex:
        logger.add('error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def add_duplicate_user_data(duplicate_data,file_path):
    uuid = file_path.split('_')[0]
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        if duplicate_data:
            existing_pkey = duplicate_data
            make_replacement(existing_pkey, duplicate_data, 'bilight_products', 'product_id',file_path)
        else:
            pass
    except Exception as _ex:
        logger.add(f'{uuid}_error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def add_products(products_data_from_user):
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.executemany(f""" INSERT INTO bilight_products (product_id,order_code,article,
                                                                manufacturer_id,product_name,tnved_id,certificate_id)
                                                                VALUES
                                                                (%s,%s,%s,%s,%s,%s,%s) """, products_data_from_user)
    except Exception as _ex:
        logger.add('error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def add_new_manufacturers(dict_to_compression, data_from_user):
    new_manufacturers = make_possible_change_dict(dict_to_compression, data_from_user)
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        for i in new_manufacturers.keys():
            if i.upper() not in dict_to_compression.keys():
                cursor.execute(f"INSERT INTO manufacturers (manufacturer_name) "
                               f"VALUES ('{i.upper()}')")
    except Exception as _ex:
        logger.add('error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def add_cert_duplicate_user_data(duplicate_data,file_path):
    uuid = file_path.split('_')[0]
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        if duplicate_data:
            existing_pkey = duplicate_data
            make_replacement(existing_pkey, duplicate_data, 'certificates', 'certificate_id',file_path)
        else:
            pass
    except Exception as _ex:
        logger.add(f'{uuid}_error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def add_certificates(cert_data_from_user,file_path):
    uuid = file_path.split('_')[0]
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.executemany(f""" INSERT INTO certificates (certificate_id,certificate_number,certificate_type_id,
                                                start_date,end_date)
                                                VALUES
                                                (%s,%s,%s,%s,%s) """, cert_data_from_user)
    except Exception as _ex:
        logger.add(f'{uuid}_error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


# endregion

# region MAKE FUNCTIONS
def make_dict(query):
    output_dict = {}
    for x in query:
        if x[1] not in output_dict:
            output_dict[x[1]] = x[0]
        else:
            pass
    return output_dict


def make_possible_change_dict(dict_to_compression, user_data_list):
    possible_change_list = []
    for i in range(len(user_data_list)):
        tmp_list = []
        editorial_number = get_editorial_num(user_data_list[i][3])
        for key in dict_to_compression.keys():
            if editorial_number >= calculate_levinstain_distance(key, user_data_list[i][3].upper()):
                tmp_list.append(key.upper())
            else:
                continue
        tmp_list.append((user_data_list[i][3]))
        possible_change_list.append(tmp_list)
    possible_change_dict = {}
    for j in range(len(possible_change_list)):
        if possible_change_list[j][-1].upper() not in dict_to_compression.keys():
            possible_change_dict[possible_change_list[j][-1]] = possible_change_list[j][:-1]
        else:
            pass
    return possible_change_dict


def make_replacement(existing_pkey, user_data, table_name, where_filter,file_path):
    uuid = file_path.split('_')[0]
    columns_name_list = get_column_name_list('information_schema.columns', table_name, 'column_name')
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        for i in range(len(existing_pkey)):
            for j in range(1, len(columns_name_list)):
                cursor.execute(f" UPDATE {table_name}"
                               f" SET {columns_name_list[j]} = %s"
                               f" WHERE {where_filter} = %s ", (user_data[i][j], user_data[i][0]))
    except Exception as _ex:
        logger.add(f'{uuid}_error.log')
        logger.error(f"{_ex}")
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")


def make_certificate_dict(query):
    output_dict = {}
    for x in query:
        if x[0] not in output_dict:
            output_dict[x[0]] = x[1:]
        else:
            pass

    return output_dict


def make_tnved_dict():
    tnved_dict = make_dict(universal_query('tnved', '*'))
    reverse_tnved_dict = dict((v, k) for k, v in tnved_dict.items())
    return reverse_tnved_dict


# endregion

# region CHECK FUNCTIONS


def approved_manufacturers_data_list(products_data_list, manufacturer_dict):
    approved_data_list = []
    for i in range(len(products_data_list)):
        if products_data_list[i][3].upper() in manufacturer_dict.keys():
            approved_data_list.append(products_data_list[i])
        else:
            pass
    return approved_data_list


def manufacturers_to_check_data_list(products_data_list):
    manufacturer_dict = make_dict(universal_query('manufacturers', '*'))
    data_list_to_check = []
    for i in range(len(products_data_list)):
        if str(products_data_list[i][3]).upper() not in manufacturer_dict:
            data_list_to_check.append(products_data_list[i])
        else:
            pass
    return data_list_to_check


# endregion

# region FIND FUNCTIONS
def find_duplicate_data(universal_query, file_name):
    data_from_user = get_product_info_from_user(file_name)
    existing_pkey = universal_query
    existing_pkey = [x[0] for x in existing_pkey]
    duplicate_user_data = []

    for i in range(len(data_from_user)):
        if data_from_user[i][0] in existing_pkey:
            duplicate_user_data.append(data_from_user[i])
        else:
            pass
    return duplicate_user_data


def find_cert_duplicate_data(universal_query, file_name):
    data_from_user = get_cert_info_from_user(file_name)
    existing_pkey = universal_query
    existing_pkey = [x[0] for x in existing_pkey]
    duplicate_user_data = []

    for i in range(len(data_from_user)):
        if data_from_user[i][0] in existing_pkey:
            duplicate_user_data.append(data_from_user[i])
        else:
            pass
    return duplicate_user_data


def find_unique_data(universal_query, data_from_user):
    existing_pkey = universal_query
    existing_pkey = [x[0] for x in existing_pkey]
    unique_user_data = []
    for i in range(len(data_from_user)):
        if data_from_user[i][0] not in existing_pkey:
            unique_user_data.append(data_from_user[i])
        else:
            pass
    return unique_user_data


def find_cert_unique_data(universal_query, data_from_user):
    existing_pkey = universal_query
    existing_pkey = [x[0] for x in existing_pkey]
    unique_user_data = []
    for i in range(len(data_from_user)):
        if data_from_user[i][0] not in existing_pkey:
            unique_user_data.append(data_from_user[i])
        else:
            pass
    return unique_user_data


# endregion

# region Make documents for user


def make_replace_file(file_path, possible_change, user_name):
    uuid_name = str(uuid.uuid4())
    uuid_dict = {}
    if uuid_name in uuid_dict.keys():
        uuid_name = str(uuid.uuid4())
    else:
        uuid_dict[uuid_name] = file_path
    extension = file_path.split('.')[-1]

    shutil.copy(rf"{file_path}", rf"possible_changes_{uuid_name}.{extension}")
    book = openpyxl.open(rf"possible_changes_{uuid_name}.{extension}", read_only=False, data_only=True)
    sheet = book.active
    uuid_file_name = rf"possible_changes_{uuid_name}.{extension}"
    font_data_to_check = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FFAA0000'
    )
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.column_dimensions['D'].width = 50
    rows_to_del_list = []
    for row in range(2, sheet.max_row + 1):
        if sheet[row][3].value not in possible_change.keys():
            rows_to_del_list.append(row)
        else:
            pass
    rows_to_del_list.sort()
    offset = 0
    for x in rows_to_del_list:
        sheet.delete_rows(idx=x - offset)
        offset += 1
    for row in range(2, sheet.max_row + 2):
        if sheet[row][3].value in possible_change.keys():
            sheet[row][3].font = font_data_to_check
            sheet[row][3].alignment = alignment
            sheet[row][3].value = f" Список замен для производителя {sheet[row][3].value}" \
                                  f" следующий:\n{possible_change[sheet[row][3].value]}\n" \
                                  f"Вставьте в ячейку производителя из предложенных и повторите загрузку\n" \
                                  f"Либо вставьте своего производителя, если уверенны в корректности названия"
    book.save(rf"possible_changes_{user_name}.{extension}")
    book.close()

    return uuid_file_name


def make_manufacturer_list_file(file_path, user_name):
    uuid_name = str(uuid.uuid4())
    uuid_dict = {}
    if uuid_name in uuid_dict.keys():
        uuid_name = str(uuid.uuid4())
    else:
        uuid_dict[uuid_name] = file_path
    extension = file_path.split('.')[-1]
    shutil.copy(rf"{file_path}", rf"manufacturer_list_by_articles_{uuid_name}.{extension}")
    book = openpyxl.open(rf"manufacturer_list_by_articles_{uuid_name}.{extension}", read_only=False, data_only=True)
    sheet = book.active
    uuid_file_name = rf"manufacturer_list_by_articles_{uuid_name}.{extension}"
    font_header = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    font_data_exist = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF005500'
    )
    font_data_not_exist = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FFAA0000'
    )
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    max_col = sheet.max_column
    sheet.cell(row=5, column=max_col + 1)
    sheet[5][max_col].value = 'Производитель'
    sheet[5][max_col].font = font_header
    sheet[5][max_col].alignment = alignment
    sheet.column_dimensions['H'].width = 20

    left_join_table_info = make_global_info_table()
    for row in range(6, sheet.max_row):
        for i in left_join_table_info:
            if sheet[row][1].value == i[0] or sheet[row][1].value == i[1]:
                sheet.cell(row=row, column=max_col + 1)
                sheet[row][max_col].value = i[2]
                sheet[row][max_col].font = font_data_exist
                break
            else:
                sheet.cell(row=row, column=max_col + 1)
                sheet[row][max_col].value = 'Нет данных'
                sheet[row][max_col].font = font_data_not_exist

    book.save(rf"manufacturer_list_by_articles_{user_name}.{extension}")
    book.close()

    return uuid_file_name


def make_certificate_list_file(file_path, user_name):
    uuid_name = str(uuid.uuid4())
    uuid_dict = {}
    if uuid_name in uuid_dict.keys():
        uuid_name = str(uuid.uuid4())
    else:
        uuid_dict[uuid_name] = file_path
    extension = file_path.split('.')[-1]
    shutil.copy(rf"{file_path}", rf"certificates_list_by_article_{uuid_name}.{extension}")
    book = openpyxl.open(rf"certificates_list_by_article_{uuid_name}.{extension}", read_only=False, data_only=True)
    sheet = book.active
    uuid_file_name = rf"certificates_list_by_article_{uuid_name}.{extension}"
    font_header = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    font_data_exist = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF005500'
    )
    font_data_not_exist = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FFAA0000'
    )
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    max_col = sheet.max_column
    for i in range(1, 5):
        for j in range(5, sheet.max_row):
            x = sheet.cell(row=j, column=max_col + i)
            x.font = font_header
            x.alignment = alignment

    sheet[5][7].value = 'Сертификат'
    sheet[5][7].font = font_header
    sheet[5][7].alignment = alignment
    sheet.column_dimensions['H'].width = 20

    sheet[5][8].value = 'Тип'
    sheet[5][8].font = font_header
    sheet[5][8].alignment = alignment

    sheet[5][9].value = 'Дата начала действия сертификата'
    sheet[5][9].font = font_header
    sheet[5][9].alignment = alignment

    sheet[5][10].value = 'Дата окончания действия сертификата'
    sheet[5][10].font = font_header
    sheet[5][10].alignment = alignment

    max_col = sheet.max_column
    left_join_table_info = make_global_info_table()
    for row in range(6, sheet.max_row):
        for i in left_join_table_info:
            if sheet[row][1].value == i[0] or sheet[row][1].value == i[1]:
                n = 3
                for x in range(4, 0, -1):
                    sheet[row][max_col - x].value = i[n]
                    sheet[row][max_col - x].font = font_data_exist
                    n += 1
                break
            else:
                for x in range(4, 0, -1):
                    sheet[row][max_col - x].value = 'Нет данных'
                    sheet[row][max_col - x].font = font_data_not_exist

    book.save(rf"certificates_list_by_article_{user_name}.{extension}")
    book.close()
    return uuid_file_name


def make_tnved_list_file(file_path, user_name):
    uuid_name = str(uuid.uuid4())
    uuid_dict = {}
    if uuid_name in uuid_dict.keys():
        uuid_name = str(uuid.uuid4())
    else:
        uuid_dict[uuid_name] = file_path
    extension = file_path.split('.')[-1]
    shutil.copy(rf"{file_path}", rf"tnved_list_by_article_{uuid_name}.{extension}")
    book = openpyxl.open(rf"tnved_list_by_article_{uuid_name}.{extension}", read_only=False, data_only=True)
    sheet = book.active
    uuid_file_name = rf"tnved_list_by_article_{uuid_name}.{extension}"
    font_header = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    font_data_exist = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF005500'
    )
    font_data_not_exist = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FFAA0000'
    )
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    max_col = sheet.max_column
    for i in range(1, 5):
        for j in range(5, sheet.max_row):
            x = sheet.cell(row=j, column=max_col + i)
            x.font = font_header
            x.alignment = alignment

    sheet[5][7].value = 'Код ТНВЭД'
    sheet[5][7].font = font_header
    sheet[5][7].alignment = alignment
    sheet.column_dimensions['H'].width = 20

    sheet[5][8].value = 'ОПИСАНИЕ'
    sheet[5][8].font = font_header
    sheet[5][8].alignment = alignment
    sheet.column_dimensions['I'].width = 150

    max_col = sheet.max_column
    left_join_table_info = make_global_info_table()

    for row in range(6, sheet.max_row):
        for i in left_join_table_info:
            if sheet[row][1].value == i[0] or sheet[row][1].value == i[1]:
                n = 7
                for x in range(4, 2, -1):
                    sheet[row][max_col - x].value = i[n]
                    sheet[row][max_col - x].font = font_data_exist
                    n += 1
                break
            else:
                for x in range(4, 2, -1):
                    sheet[row][max_col - x].value = 'Нет данных'
                    sheet[row][max_col - x].font = font_data_not_exist
    book.save(rf"tnved_list_by_article_{user_name}.{extension}")
    book.close()
    return uuid_file_name


def make_total_list_file(file_path, user_name):
    uuid_name = str(uuid.uuid4())
    uuid_dict = {}
    if uuid_name in uuid_dict.keys():
        uuid_name = str(uuid.uuid4())
    else:
        uuid_dict[uuid_name] = file_path
    extension = file_path.split('.')[-1]
    shutil.copy(rf"{file_path}", rf"total_list_by_article_{uuid_name}.{extension}")
    book = openpyxl.open(rf"total_list_by_article_{uuid_name}.{extension}", read_only=False, data_only=True)
    sheet = book.active
    uuid_file_name = rf"total_list_by_article_{uuid_name}.{extension}"
    font_header = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
    )
    font_data_exist = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF005500'
    )
    font_data_not_exist = Font(
        name='Tahoma',
        size=9,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FFAA0000'
    )
    alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    max_col = sheet.max_column
    for i in range(1, 9):
        for j in range(5, sheet.max_row):
            x = sheet.cell(row=j, column=max_col + i)
            x.font = font_header
            x.alignment = alignment

    sheet[5][7].value = 'Производитель'
    sheet[5][7].font = font_header
    sheet[5][7].alignment = alignment

    sheet[5][8].value = 'Сертификат'
    sheet[5][8].font = font_header
    sheet[5][8].alignment = alignment
    sheet.column_dimensions['H'].width = 20

    sheet[5][9].value = 'Тип'
    sheet[5][9].font = font_header
    sheet[5][9].alignment = alignment

    sheet[5][10].value = 'Дата начала действия сертификата'
    sheet[5][10].font = font_header
    sheet[5][10].alignment = alignment

    sheet[5][11].value = 'Дата окончания действия сертификата'
    sheet[5][11].font = font_header
    sheet[5][11].alignment = alignment

    sheet[5][12].value = 'ТНВЭД КОД'
    sheet[5][12].font = font_header
    sheet[5][12].alignment = alignment

    sheet[5][13].value = 'ТНВЭД ОПИСАНИЕ'
    sheet[5][13].font = font_header
    sheet[5][13].alignment = alignment

    max_col = sheet.max_column
    left_join_table_info = make_global_info_table()

    for row in range(6, sheet.max_row):
        for i in left_join_table_info:
            if sheet[row][1].value == i[0] or sheet[row][1].value == i[1]:
                n = 2
                for x in range(8, 1, -1):
                    sheet[row][max_col - x].value = i[n]
                    sheet[row][max_col - x].font = font_data_exist
                    n += 1
                break
            else:
                for x in range(8, 1, -1):
                    sheet[row][max_col - x].value = 'Нет данных'
                    sheet[row][max_col - x].font = font_data_not_exist

    book.save(rf"total_list_by_article_{user_name}.{extension}")
    book.close()
    return uuid_file_name


# endregion

# region DECODE OR CONVERT FUNCTIONS
def convert_manufacturers_to_digit(data):
    manufacturers_dict = make_dict(universal_query('manufacturers', '*'))
    for i in range(len(data)):
        if str(data[i][3]).upper() in manufacturers_dict.keys():
            data[i][3] = manufacturers_dict[data[i][3].upper()]
    return data


# def decode_certificate_types():
#     decode_dict = make_dict(universal_query('certificates_types', '*'))
#     reverse_decode_dict = dict((v, k) for k, v in decode_dict.items())
#     return reverse_decode_dict


def make_global_info_table():
    try:
        connect = psycopg2.connect(dbname=os.getenv('db_name'), user=os.getenv('user'),
                                   password=os.getenv('password'), host=os.getenv('host'))
        connect.autocommit = True
        cursor = connect.cursor()
        cursor.execute(
            f"SELECT product_id,order_code, manufacturer_name,certificate_number,"
            f"certificate_type,start_date,end_date,tnved_id,tnved_description"
            f" FROM bilight_products"
            f" LEFT JOIN manufacturers USING (manufacturer_id)"
            f" LEFT JOIN certificates USING (certificate_id)"
            f" LEFT JOIN certificates_types USING (certificate_type_id)"
            f" LEFT JOIN tnved USING (tnved_id)")
        query_data = cursor.fetchmany(9)
        return query_data
    except Exception as _ex:
        logger.add('error.log')
        logger.error(f'{_ex}')
        logger.remove()
    finally:
        cursor.close()
        connect.close()
        logger.info("Postgre SQL Connection closed")

# endregion
